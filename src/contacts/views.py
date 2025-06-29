"""
ViewSets for the contacts app.

Provides REST API endpoints for contacts, activities, and lists
with cursor pagination, filtering, and bulk operations.
"""

from django.db.models import Q, Count, Prefetch
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets
from platform_core.core.views import PlatformViewSet, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.pagination import CursorPagination
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import OrderingFilter

import csv
import io
from openpyxl import Workbook

from .models import Contact, ContactActivity, ContactList, ContactPartner, RelationshipType
from .serializers import (
    ContactSerializer, ContactActivitySerializer, ContactListSerializer,
    ContactListDetailSerializer, ContactImportSerializer, ContactExportSerializer,
    ContactPartnerSerializer
)
from .filters import ContactFilter, ContactActivityFilter, ContactListFilter


class GroupFilteredModelMixin:
    """Mixin to filter querysets by user's group for multi-tenant support."""
    
    def get_queryset(self):
        """Filter queryset by user's group."""
        # Get the model's default queryset
        if hasattr(self, 'model'):
            queryset = self.model.objects.all()
        else:
            queryset = super().get_queryset()
        
        # Filter by user's group if authenticated
        if self.request.user.is_authenticated:
            user_groups = self.request.user.groups.all()
            if user_groups.exists():
                queryset = queryset.filter(group__in=user_groups)
            else:
                # No groups, return empty queryset
                queryset = queryset.none()
        else:
            # Not authenticated, return empty queryset
            queryset = queryset.none()
        
        return queryset


class ContactCursorPagination(CursorPagination):
    """Cursor pagination for contacts ordered by creation date."""
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 200
    ordering = '-created_at'


class ContactViewSet(GroupFilteredModelMixin, PlatformViewSet):
    """
    ViewSet for Contact model with full CRUD operations.
    
    Supports:
    - Cursor pagination for efficient large dataset handling
    - Advanced filtering and search
    - Bulk operations (import/export)
    - Activity timeline
    - Lead scoring updates
    """
    
    model = Contact
    serializer_class = ContactSerializer
    pagination_class = ContactCursorPagination
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = ContactFilter
    ordering_fields = ['created_at', 'updated_at', 'last_activity_at', 'current_score']
    ordering = ['-created_at']
    
    def get_queryset(self):
        """Get contacts filtered by user's group with optimized queries."""
        queryset = super().get_queryset()
        
        # Optimize queries with select_related and prefetch_related
        queryset = queryset.select_related('assigned_to', 'group')
        
        # For list view, exclude activities to improve performance
        if self.action == 'list':
            self.get_serializer_context()['exclude_activities'] = True
        else:
            # For detail view, prefetch activities
            queryset = queryset.prefetch_related(
                Prefetch(
                    'activities',
                    queryset=ContactActivity.objects.select_related('actor').order_by('-created_at')[:5]
                ),
                'partner_relationships__partner'
            )
        
        return queryset
    
    @action(detail=False, methods=['post'])
    def bulk_import(self, request):
        """
        Bulk import contacts from CSV/Excel data.
        
        Expects JSON payload with contact data.
        """
        serializer = ContactImportSerializer(
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        
        result = serializer.save()
        
        return Response({
            'status': 'success',
            'created': result['created'],
            'updated': result['updated'],
            'skipped': result['skipped'],
            'skipped_emails': result['skipped_emails']
        })
    
    @action(detail=False, methods=['post'])
    def export(self, request):
        """
        Export contacts to CSV or Excel format.
        
        Supports filtering and field selection.
        """
        serializer = ContactExportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())
        
        # Apply additional filters from request
        if serializer.validated_data.get('list_id'):
            queryset = queryset.filter(
                contact_lists__id=serializer.validated_data['list_id']
            )
        
        # Determine fields to export
        fields = serializer.validated_data.get('fields', [
            'email', 'first_name', 'last_name', 'company_name',
            'contact_type', 'status', 'current_score', 'city', 'country',
            'created_at'
        ])
        
        if serializer.validated_data['format'] == 'csv':
            return self._export_csv(queryset, fields)
        else:
            return self._export_excel(queryset, fields)
    
    def _export_csv(self, queryset, fields):
        """Export contacts as CSV."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="contacts_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.DictWriter(response, fieldnames=fields)
        writer.writeheader()
        
        for contact in queryset.values(*fields):
            writer.writerow(contact)
        
        return response
    
    def _export_excel(self, queryset, fields):
        """Export contacts as Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"
        
        # Write headers
        for col, field in enumerate(fields, 1):
            ws.cell(row=1, column=col, value=field.replace('_', ' ').title())
        
        # Write data
        for row, contact in enumerate(queryset.values(*fields), 2):
            for col, field in enumerate(fields, 1):
                value = contact.get(field, '')
                # Convert datetime to string for Excel
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=row, column=col, value=value)
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="contacts_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    
    @action(detail=True, methods=['post'])
    def calculate_score(self, request, pk=None):
        """
        Recalculate and update the lead score for a contact.
        
        Returns the new score.
        """
        contact = self.get_object()
        new_score = contact.calculate_score()
        contact.current_score = new_score
        contact.save(update_fields=['current_score'])
        
        return Response({
            'id': contact.id,
            'current_score': new_score
        })
    
    @action(detail=True, methods=['get'])
    def activities(self, request, pk=None):
        """
        Get paginated activity timeline for a contact.
        
        Supports filtering by activity type and date range.
        """
        contact = self.get_object()
        activities = contact.activities.select_related('actor').order_by('-created_at')
        
        # Apply activity filters
        filterset = ContactActivityFilter(request.GET, queryset=activities)
        activities = filterset.qs
        
        # Paginate
        page = self.paginate_queryset(activities)
        if page is not None:
            serializer = ContactActivitySerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = ContactActivitySerializer(activities, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def add_activity(self, request, pk=None):
        """Add a new activity to a contact."""
        contact = self.get_object()
        serializer = ContactActivitySerializer(
            data=request.data,
            context={'request': request, 'contact': contact}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    @action(detail=False, methods=['get'])
    def search(self, request):
        """
        Full-text search endpoint for contacts.
        
        Uses PostgreSQL full-text search for performance.
        """
        query = request.GET.get('q', '')
        if not query:
            return Response({'results': []})
        
        # Use the search filter without get_queryset() to avoid prefetch issues
        user_groups = self.request.user.groups.all()
        if user_groups.exists():
            queryset = Contact.objects.filter(group__in=user_groups)
        else:
            queryset = Contact.objects.none()
            
        filter_instance = ContactFilter(
            {'search': query},
            queryset=queryset
        )
        
        # Get results without prefetch_related for search
        results = filter_instance.qs[:20]  # Limit to 20 results
        
        # Force evaluation before serialization
        results_list = list(results)
        
        serializer = ContactSerializer(
            results_list,
            many=True,
            context={'exclude_activities': True}
        )
        
        return Response({
            'query': query,
            'count': len(results_list),
            'results': serializer.data
        })
    
    @action(detail=False, methods=['post'])
    def assign_to_partner(self, request):
        """Bulk assign contacts to a development partner."""
        contact_ids = request.data.get('contact_ids', [])
        partner_id = request.data.get('partner_id')
        relationship_type = request.data.get('relationship_type', RelationshipType.SECONDARY)
        
        if not contact_ids or not partner_id:
            return Response(
                {'error': 'contact_ids and partner_id are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Verify partner exists and user has access
        from assessments.models import DevelopmentPartner
        try:
            partner = DevelopmentPartner.objects.filter(
                id=partner_id,
                group__in=request.user.groups.all()
            ).first()
            if not partner:
                raise DevelopmentPartner.DoesNotExist
        except DevelopmentPartner.DoesNotExist:
            return Response(
                {'error': 'Partner not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get contacts
        contacts = Contact.objects.filter(
            id__in=contact_ids,
            group__in=request.user.groups.all()
        )
        
        # Create relationships
        created = 0
        for contact in contacts:
            _, is_created = ContactPartner.objects.get_or_create(
                contact=contact,
                partner=partner,
                defaults={
                    'group': request.user.groups.first(),
                    'relationship_type': relationship_type
                }
            )
            if is_created:
                created += 1
        
        return Response({
            'status': 'success',
            'assigned': created,
            'already_assigned': len(contact_ids) - created
        })
    
    @action(detail=False, methods=['post'])
    def remove_from_partner(self, request):
        """Bulk remove contacts from a development partner."""
        contact_ids = request.data.get('contact_ids', [])
        partner_id = request.data.get('partner_id')
        
        if not contact_ids or not partner_id:
            return Response(
                {'error': 'contact_ids and partner_id are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Delete relationships
        deleted = ContactPartner.objects.filter(
            contact_id__in=contact_ids,
            partner_id=partner_id,
            group__in=request.user.groups.all()
        ).delete()[0]
        
        return Response({
            'status': 'success',
            'removed': deleted
        })


class ContactActivityViewSet(GroupFilteredModelMixin, PlatformViewSet):
    """
    ViewSet for ContactActivity model.
    
    Provides activity tracking and timeline views.
    """
    
    model = ContactActivity
    serializer_class = ContactActivitySerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = ContactActivityFilter
    ordering = ['-created_at']
    
    def get_queryset(self):
        """Get activities with optimized queries."""
        queryset = super().get_queryset()
        return queryset.select_related('contact', 'actor', 'content_type')
    
    def perform_create(self, serializer):
        """Set actor to current user when creating activity."""
        serializer.save(
            actor=self.request.user,
            group=self.request.user.groups.first()
        )


class ContactListViewSet(GroupFilteredModelMixin, PlatformViewSet):
    """
    ViewSet for ContactList model.
    
    Manages static and dynamic contact lists/segments.
    """
    
    model = ContactList
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = ContactListFilter
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        """Use detailed serializer for retrieve action."""
        if self.action == 'retrieve':
            return ContactListDetailSerializer
        return ContactListSerializer
    
    def get_queryset(self):
        """Get lists with contact counts."""
        queryset = super().get_queryset()
        
        # Annotate with contact count
        queryset = queryset.annotate(
            contact_count=Count('contacts')
        )
        
        # For detail view, prefetch contacts
        if self.action == 'retrieve':
            queryset = queryset.prefetch_related(
                'contacts__assigned_to',
                'contacts__partner_relationships'
            )
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def add_contacts(self, request, pk=None):
        """Add contacts to a list by IDs."""
        contact_list = self.get_object()
        contact_ids = request.data.get('contact_ids', [])
        
        if not contact_ids:
            return Response(
                {'error': 'contact_ids is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get contacts in the same group
        contacts = Contact.objects.filter(
            id__in=contact_ids,
            group=request.user.groups.first()
        )
        
        # Add to list
        contact_list.contacts.add(*contacts)
        
        return Response({
            'status': 'success',
            'added': contacts.count()
        })
    
    @action(detail=True, methods=['post'])
    def remove_contacts(self, request, pk=None):
        """Remove contacts from a list by IDs."""
        contact_list = self.get_object()
        contact_ids = request.data.get('contact_ids', [])
        
        if not contact_ids:
            return Response(
                {'error': 'contact_ids is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Remove from list
        contact_list.contacts.remove(*contact_ids)
        
        return Response({
            'status': 'success',
            'removed': len(contact_ids)
        })
    
    @action(detail=True, methods=['post'])
    def refresh(self, request, pk=None):
        """Refresh a dynamic list based on its filter criteria."""
        contact_list = self.get_object()
        
        if not contact_list.is_dynamic:
            return Response(
                {'error': 'Only dynamic lists can be refreshed'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # TODO: Implement dynamic list refresh logic
        # This would parse filter_criteria and update contacts
        
        return Response({
            'status': 'success',
            'message': 'Dynamic list refresh not yet implemented'
        })


class ContactPartnerViewSet(GroupFilteredModelMixin, PlatformViewSet):
    """
    ViewSet for ContactPartner relationships.
    
    Manages relationships between contacts and development partners.
    """
    
    model = ContactPartner
    serializer_class = ContactPartnerSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    ordering = ['-created_at']
    
    def get_queryset(self):
        """Get relationships with related data."""
        queryset = super().get_queryset()
        return queryset.select_related('contact', 'partner')
    
    def perform_create(self, serializer):
        """Ensure group assignment when creating relationship."""
        serializer.save(group=self.request.user.groups.first())
